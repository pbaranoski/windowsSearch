import tkinter as tk
from tkinter import filedialog as fd
from tkinter import messagebox
#from tkinter.tix import *

import pandas as pd
import openpyxl as excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Fill, Color

import os

cur_dir = ""
sFileHeader = ""
bChoicesLoaded = False

TBL_COLS1_csv = ""
TBL_COLS2_csv = ""

#fDtlDiffs = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\PBAR_IMPL\IMPL_CD_SGNTRS_Diffs.csv"
fDtlDiffsXLSX = ""


def get_csv_file(lblLabel):

    global cur_dir
    global sFileHeader
    global bChoicesLoaded

    filetypes = (
        ('csv files', '*.csv'),
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir=cur_dir,
        filetypes=filetypes 
    )

    # Verify filename has extension
    arrFileParts = os.path.splitext(filename)
    if arrFileParts[1].strip() == "":
        filename = arrFileParts[0] + ".csv"

    # Assign selected filename to screen label
    lblLabel.config(text=filename)

    # set the current default directory for dialog box
    cur_dir = os.path.dirname(filename)
    print (cur_dir)

    #################################################################
    # Read header record for file to get list of omit column choices
    #################################################################
    if bChoicesLoaded == False:
        bChoicesLoaded = True
        with open(filename, "r", encoding = 'utf-8') as f:
            sFileHeader = f.readline()
            print("header: "+sFileHeader)

        loadOmitChoices()    


def loadOmitChoices():

    global tkDropdownSelection
    # Create a Tkinter variable
    tkDropdownSelection = tk.StringVar(MDIWnd)

    # Dictionary with options
    lstOmitChoices = sFileHeader.split(",")
    tkDropdownSelection.set(lstOmitChoices[0]) # set the default option

    popupMenu = tk.OptionMenu(MDIWnd, tkDropdownSelection, *lstOmitChoices, command=dropdownSelectionMade)
    popupMenu.grid(row=4, column=4, pady=5)


def dropdownSelectionMade(choice):

    global sDropDownSelections
    choice = tkDropdownSelection.get()
    #print(f"choice:{choice}")

    # Get current textbox text 
    sDropDownSelections = txtOmitCols.get()

    # Add selection to textbox
    if sDropDownSelections == "":
        sDropDownSelections = choice
    else:     
        sDropDownSelections += ", " + choice 

    # modify text in textbox to include new selection
    txtOmitCols.delete(0, 'end')
    txtOmitCols.insert(0, sDropDownSelections)

    print(f"sDropDownSelections: {sDropDownSelections}")


def sel_xlsx_file(lblLabel):

    global cur_dir

    filetypes = (
        ('Excel files', '*.xlsx'),
    )

    filename = fd.asksaveasfilename(
        title='SaveAs file',
        initialdir=cur_dir,
        filetypes=filetypes 
    )

    # Verify filename has extension
    # and that extension is xlsx
    arrFileParts = os.path.splitext(filename)
    if arrFileParts[1].strip() == "":
        filename = arrFileParts[0] + ".xlsx"
    elif arrFileParts[1].strip() != "xlsx":
        filename = arrFileParts[0] + ".xlsx"     

    # Assign selected filename to screen label
    lblLabel.config(text=filename)

    # set the current default directory for dialog box
    cur_dir = os.path.dirname(filename)
    print (cur_dir)


def initiateCompareFilesAction():

    global TBL_COLS1_csv
    global TBL_COLS2_csv
    global fDtlDiffsXLSX
    global lstOmitColumns

    ########################################
    # 1) Perform validation of input fields
    # 2) Perform Compare
    ########################################

    # load variables with window values
    txtInFile1 = lblInFile1Text.cget("text")
    txtInFile2 = lblInFile2Text.cget("text")
    txtOutFile = lblOutFileText.cget("text")

    # Verify that input values have been entered/selected

    if txtInFile1 == "":
        messagebox.showerror("Error", "Input File 1 has not been selected!")
        return

    if txtInFile2 == "":
        messagebox.showerror("Error", "Input File 2 has not been selected!")
        return

    if txtOutFile == "":
        messagebox.showerror("Error", "Output File has not been selected!")
        return

    #####################################################
    # 1) Load variables needed by compareFiles function
    #    using screen input values.
    # 2) Perform csv file comparison
    #####################################################
    TBL_COLS1_csv = txtInFile1
    TBL_COLS2_csv = txtInFile2
    fDtlDiffsXLSX = txtOutFile
    #sColumns = txtOmitCols.get().replace(" ","")
    sColumns = txtOmitCols.get().replace(",,",",")

    lstOmitColumns = sColumns.split(",")
    for i in range(0,len(lstOmitColumns)):
        lstOmitColumns[i] = lstOmitColumns[i].strip()

    print(lstOmitColumns)
    compareFiles()

    ###########################################
    # Comparison has completed
    ###########################################
    messagebox.showinfo("Complete", "Results have been generated!")


def removeDFCols(df, lstCols2Remove):
    # sColTxt = Column to drop from Data Frame

    lstCols = df.columns.values.tolist()

    for sColTxt in lstCols2Remove:
        idx = lstCols.count(sColTxt)
        if idx > 0:
            df.drop(sColTxt, axis=1, inplace=True)

    return df    


def removeIDRNulls(df):

    lstCols = df.columns.values.tolist()
    for col in lstCols:
        df.loc[df[col] == "?", col] = '' 

    return df


def trimTrailingSpaces(df):
    
    # Remove Leading and Trailing spaces in Pandas cells    
    df.replace({"^\s*|\s*$":""}, regex=True, inplace=True) 

    """    
    lstCols = df.columns.values.tolist()

    for col in lstCols:
        if col == "CLM_LINE_INVLD_HCPCS_CD":
            print( "|"+ df['CLM_LINE_INVLD_HCPCS_CD'].iloc[0] + "|")
    """

    return df


def sortDF(df):

    lstCols = df.columns.values.tolist()
    #df = df.sort_values(by=['TABLE_SCHEMA', 'TABLE_NAME', 'COLUMN_NAME'])
    df = df.sort_values(by=lstCols)
    
    return df 


def compareFiles():
    #########################################################
    # Define variables
    #########################################################

    # truncate output file if exists
    ##if os.path.exists(fDtlDiffs):
    ##    os.truncate(fDtlDiffs, 0)

    #########################################################
    # Read CME and IDR csv files into Pandas Data Frames
    #########################################################
    dfFile1 = pd.read_csv(TBL_COLS1_csv, dtype=str, na_filter=False) 
    dfFile1 = removeDFCols(dfFile1, ["IDR_INSRT_TS", "IDR_UPDT_TS"]) 
    dfFile1 = removeDFCols(dfFile1, lstOmitColumns)     
    dfFile1 = removeIDRNulls(dfFile1)
    dfFile1 = trimTrailingSpaces(dfFile1)

    dfFile2 = pd.read_csv(TBL_COLS2_csv, dtype=str, na_filter=False) 
    dfFile2 = removeDFCols(dfFile2, ["IDR_INSRT_TS", "IDR_UPDT_TS"]) 
    dfFile2 = removeDFCols(dfFile2, lstOmitColumns)   
    dfFile2 = removeIDRNulls(dfFile2)
    dfFile2 = trimTrailingSpaces(dfFile2)

    print("NOF dfFile1 rows:"+str(len(dfFile1.index)))
    print("NOF dfFile2 rows:"+str(len(dfFile2.index)))

    #########################################################################################
    # In essence a join/comparison on all columns
    # NOTE: Setting "indicator=True" adds a column to the merged DataFame where the value 
    #       of each row can be one of three possible values: left_only, right_only, or both
    # NOTE2: Another option "first_df.compare(second_df, keep_equal=True)"
    #########################################################################################
    comparison_df = dfFile1.merge(dfFile2, indicator=True, how='outer')
    comparison_df = sortDF(comparison_df)

    # All Comparisons
    ##comparison_df.to_csv(fDtlDiffs,sep = ",", index=False, line_terminator = "\n")  

    # after merge --> row will be "marked" as 'both" if both data frame rows match
    dfDiff = comparison_df[comparison_df['_merge'] != 'both']
    dfDiff = sortDF(dfDiff)

    dfMatches = comparison_df[comparison_df['_merge'] == 'both']
    dfMatches = sortDF(dfMatches)

    ##########################################
    # saving Results to xlsx file
    ##########################################
    XLSXWriter = pd.ExcelWriter(fDtlDiffsXLSX)

    dfDiff.to_excel(XLSXWriter, sheet_name="Differences", index=False)   
    dfMatches.to_excel(XLSXWriter, sheet_name="Matches", index=False)   

    XLSXWriter.save()


    ################################
    # Modify Excel spreadsheet
    # Add hilighting of rows
    ################################
    wrkbk = excel.load_workbook(fDtlDiffsXLSX)

    if wrkbk is None:
        print("couldn't get workbook onject")
    else:
        print("Got the workbook")    

    ###################################################
    # process Excel spreadsheet
    ###################################################
    for sheet in wrkbk:

        sheet.freeze_panes = 'A2'

        # Set Excel cell color values
        cellColorFillHdr = PatternFill(start_color='01B0F1', end_color='01B0F1', fill_type = "solid") 
        cellColorFill1 = PatternFill(start_color="DDEBF6", end_color="DDEBF6", fill_type = "solid")
        cellColorFill2 = PatternFill(start_color="E2EFDB", end_color="E2EFDB", fill_type = "solid")

        # Set Excel border value
        thin = Side(border_style="thin")
        cellBorder = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Set Alternating color pattern
        iNOFRowsWithColor = 0       
        if sheet.title == "Matches":
            iNOFRowsWithColor = 1
        else:
            iNOFRowsWithColor = 2            

        # Color switch --> used in XOR operation
        iFlipColor = 1


        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

            # Set indicator for 1st header Row
            if row[0].row == 1:
                bHeaderRow = True
            else:
                bHeaderRow = False

            # Set Row Color
            if bHeaderRow:
                # skip header
                pass
            elif (row[0].row) % iNOFRowsWithColor == 0:
                # Time to change colors
                # XOR value --> to flip color switch back-n-forth; determine which color to use
                iFlipColor ^= 1

            # Set cell attributes
            for cell in row:
                cell.border = cellBorder
                
                if bHeaderRow:
                    cell.fill = cellColorFillHdr
                else:
                    if iFlipColor == 0:
                        cell.fill = cellColorFill1
                    else:
                        cell.fill = cellColorFill2 

    # Save workbook changes                                            
    wrkbk.save(fDtlDiffsXLSX)
  
    return 0

def main():

    ####################################################
    # define variables
    ####################################################
    global MDIWnd
    global lblInFile1Text
    global lblInFile2Text
    global lblOutFileText
    global txtOmitCols

    ####################################################
    # Build window
    ####################################################
    MDIWnd= tk.Tk()
    MDIWnd.title("Compare Two CSV Files")
    MDIWnd.geometry("1000x300")

    lblHdr = tk.Label(MDIWnd, text='Compare Two CSV Files')
    lblHdr.config(font=('helvetica', 14))
    lblHdr.grid(row=0, column=3, columnspan=3, padx=5, pady=10)

    lblSpacer = tk.Label(MDIWnd, text="          ")
    lblSpacer.grid(row=0, column=0)

    ##############################
    # Select InFile1
    ##############################
    btnInFile1 = tk.Button(text='Select File', command=lambda:get_csv_file(lblInFile1Text), bg='blue', fg='white', font=('helvetica', 8, 'bold'))
    btnInFile1.grid(row=1, column=1, pady=3)

    #lblInFile1Label = tk.Label(MDIWnd, text='Input File 1:', bd=1, relief="sunken")
    lblInFile1Label = tk.Label(MDIWnd, text='Input File 1:')
    lblInFile1Label.config(font=('helvetica', 10))
    lblInFile1Label.grid(row=1, column=2, sticky='e', pady=1)

    # Selected InFile1
    lblInFile1Text = tk.Label(MDIWnd, text="")
    lblInFile1Text.config(font=('helvetica', 10))
    lblInFile1Text.grid(row=1, column=3, sticky='w')

    ###############################
    # Select InFile2
    ###############################
    btnInFile2 = tk.Button(text='Select File', command=lambda:get_csv_file(lblInFile2Text), bg='blue', fg='white', font=('helvetica', 8, 'bold'))
    btnInFile2.grid(row=2, column=1, pady=3)

    lblInFile2Label = tk.Label(MDIWnd, text='Input File 2:')
    lblInFile2Label.config(font=('helvetica', 10))
    lblInFile2Label.grid(row=2, column=2, sticky='e')

    lblInFile2Text = tk.Label(MDIWnd, text="")
    lblInFile2Text.config(font=('helvetica', 10))
    lblInFile2Text.grid(row=2, column=3, sticky='w')

    ###############################
    # Output File Label
    ###############################
    btnOutFile = tk.Button(text='Select File', command=lambda:sel_xlsx_file(lblOutFileText), bg='blue', fg='white', font=('helvetica', 8, 'bold'))
    btnOutFile.grid(row=3, column=1, padx=2, pady=3)

    lblOutFileLabel = tk.Label(MDIWnd, text='Output File:')
    lblOutFileLabel.config(font=('helvetica', 10))
    lblOutFileLabel.grid(row=3, column=2, sticky='e')

    # Selected Output File
    lblOutFileText = tk.Label(MDIWnd, text='')
    lblOutFileText.config(font=('helvetica', 10))
    lblOutFileText.grid(row=3, column=3, sticky='w')

    ###############################
    # Columns to Omit
    ###############################
    lblOmitCols = tk.Label(MDIWnd, text='Columns to Omit:')
    lblOmitCols.config(font=('helvetica', 10))
    lblOmitCols.grid(row=4, column=2, sticky='e', pady=3)

    txtOmitCols = tk.Entry (MDIWnd, width=90, justify="left") 
    txtOmitCols.grid(row=4, column=3, sticky='w')

    ###############################
    # Buttons
    ###############################
    btnCompare = tk.Button(text='Compare Files', command=initiateCompareFilesAction, bg='blue', fg='white', font=('helvetica',10, 'bold'))
    btnCompare.grid(row=6, column=3, sticky='w', pady=10)

    ####################################
    # Process Window messages
    ####################################
    MDIWnd.mainloop()


if __name__ == "__main__":
    
    main()




