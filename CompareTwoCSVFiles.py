import pandas as pd
import openpyxl as excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Fill, Color
import os

#from pymysql import ROWID
import time

#TBL_COLS1_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\BIA_DEV_INFO.csv"
#TBL_COLS2_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\BIA_TST_INFO.csv"

#TBL_COLS1_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\Snowflake\TST_CLM_CYQ_SGNTR_TEMP.csv"
#TBL_COLS2_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\Teradata\TST_V1_CLM_CYQ_SGNTR.csv"

#TBL_COLS1_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\Snowflake\TST_CLM_PTB_PROC_CNTL.csv"
#TBL_COLS2_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\Teradata\TST_V1_CLM_PTB_PROC_CNTL.csv"

TBL_COLS1_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\PBAR_IMPL\IMPL_CLM_PTB_MO_AGG.csv"
TBL_COLS2_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\PBAR_IMPL\INT_V1_CLM_PTB_MO_AGG.csv"

#fDtlDiffs = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\TBL_COL_Diffs.csv"
#fDtlDiffsXLSX = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\TBL_COL_Diffs.xlsx"
#fDtlDiffs = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\CLM_CYQ_SGNTR_Diffs.csv"
#fDtlDiffsXLSX = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\CLM_CYQ_SGNTR_Diffs.xlsx"

fDtlDiffs = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\PBAR_IMPL\IMPL_CLM_PTB_MO_AGG_DiffsTest.csv"
fDtlDiffsXLSX = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\PBAR_IMPL\IMPL_CLM_PTB_MO_AGG_DiffsTest.xlsx"


def removeDFCols(df, lstCols2Remove):
    # sColTxt = Column to drop from Data Frame

    lstCols = df.columns.values.tolist()

    for sColTxt in lstCols2Remove:
        idx = lstCols.count(sColTxt)
        if idx > 0:
            df.drop(sColTxt, axis=1, inplace=True)

    return df    


def removeIDRNulls(df):

    lstCols = df.columns.values.tolist()
    for col in lstCols:
        df.loc[df[col] == "?", col] = '' 

    return df


def trimTrailingSpaces(df):
    
    # Remove Leading and Trailing spaces in Pandas cells    
    df.replace({"^\s*|\s*$":""}, regex=True, inplace=True) 

    """    
    lstCols = df.columns.values.tolist()

    for col in lstCols:
        if col == "CLM_LINE_INVLD_HCPCS_CD":
            print( "|"+ df['CLM_LINE_INVLD_HCPCS_CD'].iloc[0] + "|")
    """

    return df


def sortDF(df):

    lstCols = df.columns.values.tolist()
    #df = df.sort_values(by=['TABLE_SCHEMA', 'TABLE_NAME', 'COLUMN_NAME'])
    df = df.sort_values(by=lstCols)
    
    return df 


def compareFiles():
    #########################################################
    # Define variables
    #########################################################

    # truncate output file if exists
    if os.path.exists(fDtlDiffs):
        os.truncate(fDtlDiffs, 0)

    #########################################################
    # Read CME and IDR csv files into Pandas Data Frames
    #########################################################
    dfFile1 = pd.read_csv(TBL_COLS1_csv, dtype=str, na_filter=False) 
    dfFile1 = removeDFCols(dfFile1, ["IDR_INSRT_TS", "IDR_UPDT_TS"]) 
    #dfFile1 = removeDFCols(dfFile1, ["CLM_CYQ_SGNTR_SK", "CLM_YEAR_SGNTR_SK"])  
    dfFile1 = removeDFCols(dfFile1, ["CLM_YEAR_SGNTR_SK", "CLM_CYQ_SGNTR_SK", "CLM_MO_SGNTR_SK", "CLM_CD_SGNTR_SK"])     
    dfFile1 = removeIDRNulls(dfFile1)
    dfFile1 = trimTrailingSpaces(dfFile1)

    dfFile2 = pd.read_csv(TBL_COLS2_csv, dtype=str, na_filter=False) 
    dfFile2 = removeDFCols(dfFile2, ["IDR_INSRT_TS", "IDR_UPDT_TS"]) 
    #dfFile2 = removeDFCols(dfFile2, ["CLM_CYQ_SGNTR_SK", "CLM_YEAR_SGNTR_SK"])     
    dfFile2 = removeDFCols(dfFile2, ["CLM_YEAR_SGNTR_SK", "CLM_CYQ_SGNTR_SK", "CLM_MO_SGNTR_SK", "CLM_CD_SGNTR_SK"])     
    dfFile2 = removeIDRNulls(dfFile2)
    dfFile2 = trimTrailingSpaces(dfFile2)

    print("NOF dfFile1 rows:"+str(len(dfFile1.index)))
    print("NOF dfFile2 rows:"+str(len(dfFile2.index)))

    #########################################################################################
    # In essence a join/comparison on all columns
    # NOTE: Setting "indicator=True" adds a column to the merged DataFame where the value 
    #       of each row can be one of three possible values: left_only, right_only, or both
    # NOTE2: Another option "first_df.compare(second_df, keep_equal=True)"
    #########################################################################################
    comparison_df = dfFile1.merge(dfFile2, indicator=True, how='outer')
    comparison_df = sortDF(comparison_df)

    # All Comparisons
    ##comparison_df.to_csv(fDtlDiffs,sep = ",", index=False, line_terminator = "\n")  

    # after merge --> row will be "marked" as 'both" if both data frame rows match
    dfDiff = comparison_df[comparison_df['_merge'] != 'both']
    dfDiff = sortDF(dfDiff)

    dfMatches = comparison_df[comparison_df['_merge'] == 'both']
    dfMatches = sortDF(dfMatches)

    print("Comparisons are complete!"+str(time.time() ) )

    ##########################################
    # saving Results to xlsx file
    ##########################################
    XLSXWriter = pd.ExcelWriter(fDtlDiffsXLSX)

    dfDiff.to_excel(XLSXWriter, sheet_name="Differences", index=False)   
    dfMatches.to_excel(XLSXWriter, sheet_name="Matches", index=False)   

    XLSXWriter.save()

    print("Excel file is written"+str(time.time() ) )

    ################################
    # Modify Excel spreadsheet
    # Add hilighting of rows
    ################################
    wrkbk = excel.load_workbook(fDtlDiffsXLSX)

    if wrkbk is None:
        print("couldn't get workbook onject")
    else:
        print("Got the workbook")    

    ###################################################
    # process Excel spreadsheet
    ###################################################
    for sheet in wrkbk:

        sheet.freeze_panes = 'A2'

        # Set Excel cell color values
        cellColorFillHdr = PatternFill(start_color='01B0F1', end_color='01B0F1', fill_type = "solid") 
        cellColorFill1 = PatternFill(start_color="DDEBF6", end_color="DDEBF6", fill_type = "solid")
        cellColorFill2 = PatternFill(start_color="E2EFDB", end_color="E2EFDB", fill_type = "solid")

        # Set Excel border value
        thin = Side(border_style="thin")
        cellBorder = Border(top=thin, left=thin, right=thin, bottom=thin)


        # Set Alternating color pattern value
        iNOFRowsWithColor = 0       
        if sheet.title == "Matches":
            iNOFRowsWithColor = 1
        else:
            iNOFRowsWithColor = 2            

        # Color switch --> used in XOR operation
        iFlipColor = 1


        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

            # Set indicator for 1st header Row
            if row[0].row == 1:
                bHeaderRow = True
            else:
                bHeaderRow = False

            # Set Row Color
            if bHeaderRow:
                # skip header
                pass
            elif (row[0].row) % iNOFRowsWithColor == 0:
                # Time to change colors
                # XOR value --> to flip color switch back-n-forth; determine which color to use
                iFlipColor ^= 1

            # Set cell attributes
            for cell in row:
                cell.border = cellBorder
                
                if bHeaderRow:
                    cell.fill = cellColorFillHdr
                else:
                    if iFlipColor == 0:
                        cell.fill = cellColorFill1
                    else:
                        cell.fill = cellColorFill2    

    # Save workbook changes                                            
    wrkbk.save(fDtlDiffsXLSX)

    print("Formatting is done"+str(time.time() ) )

    ##exit(0)


if __name__ == "__main__":
    
    compareFiles()

